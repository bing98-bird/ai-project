import os
import pandas as pd
import numpy as np
import fitz  # PyMuPDF
from typing import Dict, List, Any
from collections import Counter
from langchain_google_genai import ChatGoogleGenerativeAI
from dotenv import load_dotenv
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

api_key = os.getenv("GEMINI_API_KEY")
llm = ChatGoogleGenerativeAI(model="gemini-3.1-flash-lite-preview", google_api_key=api_key)


def _detect_outliers(series: pd.Series) -> Dict[str, Any]:
    """Detect outliers using IQR method."""
    if series.isna().all():
        return {}
    
    Q1 = series.quantile(0.25)
    Q3 = series.quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    
    outliers = series[(series < lower_bound) | (series > upper_bound)]
    if len(outliers) > 0:
        return {
            "count": len(outliers),
            "percentage": round((len(outliers) / len(series)) * 100, 2),
            "values": outliers.head(5).tolist(),
        }
    return {}


def _get_advanced_stats(df: pd.DataFrame) -> Dict[str, Any]:
    """Extract advanced statistics from a dataframe."""
    stats = {
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "column_names": list(df.columns),
        "data_types": df.dtypes.astype(str).to_dict(),
        "missing_values": df.isnull().sum().to_dict(),
        "missing_percentage": {col: round((df[col].isnull().sum() / len(df)) * 100, 2) for col in df.columns},
        "duplicates": len(df) - len(df.drop_duplicates()),
    }
    
    # Data quality score
    total_cells = len(df) * len(df.columns)
    missing_cells = df.isnull().sum().sum()
    quality_score = round(((total_cells - missing_cells) / total_cells) * 100, 2)
    stats["data_quality_score"] = quality_score
    
    # Add numeric statistics for numeric columns
    numeric_cols = df.select_dtypes(include=["number"]).columns
    numeric_stats = {}
    for col in numeric_cols:
        if not df[col].isna().all():
            col_stats = {
                "mean": round(float(df[col].mean()), 2),
                "median": round(float(df[col].median()), 2),
                "min": round(float(df[col].min()), 2),
                "max": round(float(df[col].max()), 2),
                "std": round(float(df[col].std()), 2),
                "q25": round(float(df[col].quantile(0.25)), 2),
                "q75": round(float(df[col].quantile(0.75)), 2),
            }
            
            # Detect outliers
            outliers = _detect_outliers(df[col])
            if outliers:
                col_stats["outliers"] = outliers
            
            numeric_stats[col] = col_stats
    
    if numeric_stats:
        stats["numeric_statistics"] = numeric_stats
    
    # Add categorical value counts for key columns
    categorical_cols = df.select_dtypes(include=["object"]).columns
    categorical_stats = {}
    for col in categorical_cols[:10]:  # Limit to first 10
        value_counts = df[col].value_counts().head(10).to_dict()
        if value_counts:
            categorical_stats[col] = {
                "value_counts": value_counts,
                "unique_values": df[col].nunique(),
                "most_common": df[col].value_counts().index[0] if len(df[col].value_counts()) > 0 else None,
            }
    
    if categorical_stats:
        stats["categorical_summary"] = categorical_stats
    
    # Detect correlations between numeric columns
    if len(numeric_cols) > 1:
        try:
            corr_matrix = df[numeric_cols].corr()
            high_correlations = []
            for i in range(len(corr_matrix.columns)):
                for j in range(i+1, len(corr_matrix.columns)):
                    corr_val = corr_matrix.iloc[i, j]
                    if abs(corr_val) > 0.7:  # Strong correlation threshold
                        high_correlations.append({
                            "var1": corr_matrix.columns[i],
                            "var2": corr_matrix.columns[j],
                            "correlation": round(float(corr_val), 3),
                        })
            if high_correlations:
                stats["strong_correlations"] = high_correlations
        except:
            pass
    
    return stats


def _analyze_csv(file_path: str) -> Dict[str, Any]:
    """Analyze a CSV file and extract statistics."""
    df = pd.read_csv(file_path)
    return _get_advanced_stats(df)


def _analyze_excel(file_path: str) -> Dict[str, Any]:
    """Analyze an Excel file and extract statistics."""
    df = pd.read_excel(file_path)
    return _get_advanced_stats(df)


def _extract_pdf_text(file_path: str) -> Dict[str, Any]:
    """Extract and analyze text content from a PDF file."""
    doc = fitz.open(file_path)
    
    total_pages = len(doc)
    full_text = ""
    text_by_page = []
    
    # Extract text from all pages
    for page_num in range(total_pages):
        page = doc[page_num]
        text = page.get_text()
        full_text += text + "\n"
        text_by_page.append({
            "page": page_num + 1,
            "text": text,
            "length": len(text)
        })
    
    # Basic statistics
    words = full_text.split()
    sentences = [s.strip() for s in full_text.split('.') if s.strip()]
    
    # Word frequency analysis
    # Remove common stop words
    stop_words = {
        'the', 'a', 'an', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for',
        'of', 'with', 'by', 'from', 'as', 'is', 'was', 'are', 'be', 'been',
        'have', 'has', 'do', 'does', 'did', 'will', 'would', 'could', 'should',
        'that', 'this', 'these', 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they',
    }
    
    significant_words = [w.lower() for w in words if w.lower() not in stop_words and len(w) > 3]
    word_freq = Counter(significant_words).most_common(15)
    
    stats = {
        "total_pages": total_pages,
        "total_words": len(words),
        "total_sentences": len(sentences),
        "average_page_length": round(len(full_text) / total_pages, 0),
        "word_frequency": dict(word_freq),
        "full_text": full_text[:3000],  # First 3000 chars for context
        "text_by_page": text_by_page,
    }
    
    return stats


def _analyze_pdf(file_path: str) -> Dict[str, Any]:
    """Analyze a PDF file and extract insights."""
    return _extract_pdf_text(file_path)


def _standardize_status(status_value: str) -> str:
    """Standardize status values to consistent format."""
    if pd.isna(status_value) or status_value == "" or status_value == "???":
        return "Unknown"
    
    status = str(status_value).strip().lower()
    
    # Map variations to standard statuses
    if any(word in status for word in ["paid", "bayar", "ok", "plu"]):
        return "Paid"
    elif any(word in status for word in ["pending", "kiv", "waiting", "jupe", "pm sgn"]):
        return "Pending"
    elif any(word in status for word in ["reject", "failed", "no payment"]):
        return "Rejected"
    elif any(word in status for word in ["cash", "petty", "inv"]):
        return "Petty Cash"
    else:
        return status.title()


def _parse_financial_value(value: str) -> float:
    """Convert financial values to numeric format with enhanced validation."""
    if pd.isna(value) or value == "" or value == "???":
        return 0.0
    
    value_str = str(value).strip().upper()
    
    # Skip non-numeric status values like "Pending QS", "Pending OS", etc.
    non_numeric_keywords = ["PENDING", "QS", "OS", "TBC", "NA", "N/A", "UNKNOWN", "REJECT", "APPROVED"]
    if any(keyword in value_str for keyword in non_numeric_keywords):
        return 0.0
    
    # Remove RM prefix, commas, and spaces
    value_str = value_str.replace("RM", "").replace(",", "").strip()
    
    # Handle abbreviations (k = 1000, m = 1,000,000)
    try:
        if "K" in value_str:
            # Extract number before K
            num_str = value_str.replace("K", "").strip()
            numeric_part = ''.join(c for c in num_str if c.isdigit() or c == '.')
            if numeric_part:
                return float(numeric_part) * 1000
            else:
                return 0.0
        elif "M" in value_str:
            # Extract number before M
            num_str = value_str.replace("M", "").strip()
            numeric_part = ''.join(c for c in num_str if c.isdigit() or c == '.')
            if numeric_part:
                return float(numeric_part) * 1_000_000
            else:
                return 0.0
        else:
            # Extract only numeric parts (digits and decimal points)
            numeric_str = ''.join(c for c in value_str if c.isdigit() or c == '.')
            if numeric_str:
                return float(numeric_str)
            else:
                return 0.0
    except (ValueError, TypeError):
        return 0.0


def _clean_financial_data(df: pd.DataFrame) -> tuple[pd.DataFrame, Dict[str, Any]]:
    """Clean messy financial data and generate summary while preserving original data."""
    df_clean = df.copy()
    
    # Identify financial columns - be more specific to avoid identifier columns
    financial_cols = [col for col in df_clean.columns if any(keyword in col.lower() 
                      for keyword in ["amt", "amount", "val", "sum", "total", "contract val", "price"])]
    
    # Identify status columns
    status_cols = [col for col in df_clean.columns if any(keyword in col.lower() 
                   for keyword in ["status", "state"])]
    
    # Add cleaned financial columns with "CLEAN_" prefix to preserve original data
    for col in financial_cols:
        # Create a cleaned version of the column name
        clean_col_name = f"CLEAN_{col}".upper().replace(" ", "_")
        df_clean[clean_col_name] = df[col].apply(_parse_financial_value)
    
    # Add cleaned status columns with "CLEAN_" prefix to preserve original data
    for col in status_cols:
        clean_col_name = f"CLEAN_{col}".upper().replace(" ", "_")
        df_clean[clean_col_name] = df[col].apply(_standardize_status)
    
    # Generate financial summary using the cleaned columns for accurate totals
    summary = {
        "total_records": len(df_clean),
        "status_breakdown": {},
        "financial_summary": {}
    }
    
    # Status breakdown from cleaned status column
    if status_cols:
        main_status_col = status_cols[0]
        clean_status_col = f"CLEAN_{main_status_col}".upper().replace(" ", "_")
        status_counts = df_clean[clean_status_col].value_counts().to_dict()
        summary["status_breakdown"] = status_counts
    
    # Financial breakdown by status from cleaned columns
    amount_cols = [col for col in financial_cols if any(keyword in col.lower() 
                   for keyword in ["amt", "amount", "claimed", "claim amt"])]
    
    if amount_cols and status_cols:
        main_amount_col = amount_cols[0]
        main_status_col = status_cols[0]
        
        clean_amount_col = f"CLEAN_{main_amount_col}".upper().replace(" ", "_")
        clean_status_col = f"CLEAN_{main_status_col}".upper().replace(" ", "_")
        
        for status in df_clean[clean_status_col].unique():
            status_data = df_clean[df_clean[clean_status_col] == status]
            total_amount = status_data[clean_amount_col].sum()
            summary["financial_summary"][status] = round(total_amount, 2)
    
    return df_clean, summary


def _format_excel_file(excel_file_path: str) -> None:
    """Format Excel file with proper column widths, headers styling, and alternating row colors."""
    try:
        workbook = load_workbook(excel_file_path)
        worksheet = workbook.active
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Light blue for alternating rows (for easier reading)
        light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Border style for all cells
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Format header row and set column widths
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Auto-calculate column width based on content
            column_letter = get_column_letter(col_idx)
            max_length = len(str(cell.value)) + 2
            
            # Check all values in column to get max length
            for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell_in_col in row:
                    try:
                        if cell_in_col.value:
                            max_length = max(max_length, len(str(cell_in_col.value)) + 2)
                    except:
                        pass
            
            # Set column width (with min and max limits for readability)
            adjusted_width = min(max_length, 40)  # Cap at 40 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply alternating row colors and borders to data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # Alternating row color
                if row_idx % 2 == 0:
                    cell.fill = light_fill
                
                # Format numbers (align right)
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
        # Set row height for header
        worksheet.row_dimensions[1].height = 25
        
        workbook.save(excel_file_path)
    except Exception as e:
        # If formatting fails, continue without formatting
        print(f"Warning: Could not format Excel file: {str(e)}")


async def generate_cleaned_data_report(file_path: str, file_name: str) -> Dict[str, Any]:
    """Clean messy Excel/CSV data and generate intelligent report with download link."""
    try:
        # Read file
        if file_name.lower().endswith(".xlsx") or file_name.lower().endswith(".xls"):
            df = pd.read_excel(file_path)
        elif file_name.lower().endswith(".csv"):
            df = pd.read_csv(file_path)
        else:
            return {"status": "error", "message": "Unsupported file type"}
        
        # Clean data
        df_clean, summary = _clean_financial_data(df)
        
        # Create cleaned filename
        base_name = os.path.splitext(file_name)[0]
        ext = os.path.splitext(file_name)[1]
        cleaned_filename = f"{base_name}_CLEANED{ext}"
        cleaned_file_path = os.path.join(os.path.dirname(file_path), cleaned_filename)
        
        # Save cleaned file
        if cleaned_filename.endswith((".xlsx", ".xls")):
            df_clean.to_excel(cleaned_file_path, index=False)
            # Format Excel file for better readability
            _format_excel_file(cleaned_file_path)
        else:
            df_clean.to_csv(cleaned_file_path, index=False)
        
        # Generate detailed report
        report_text = f"""
**DATA CLEANING REPORT: {file_name}**

ORIGINAL DATA:
- Total Records: {len(df)}
- Total Columns: {len(df.columns)}
- Missing Values: {df.isnull().sum().sum()}

CLEANED DATA:
- Total Records: {len(df_clean)}
- Total Columns: {len(df_clean.columns)}
- Standardized Status Values: Yes
- Cleaned Financial Values: Yes

CLEANING ACTIONS PERFORMED:
1. ✅ Standardized status values (Paid, Pending, Rejected, Petty Cash)
2. ✅ Converted financial values to numeric format (handled RM, k, m abbreviations)
3. ✅ Handled missing and invalid data entries
4. ✅ Removed currency symbols and formatting inconsistencies
5. ✅ Consolidated duplicate/similar status entries

STATUS BREAKDOWN:
"""
        
        for status, count in summary["status_breakdown"].items():
            report_text += f"  • {status}: {count} records\n"
        
        report_text += "\nFINANCIAL SUMMARY BY STATUS:\n"
        
        total_paid = sum(v for k, v in summary["financial_summary"].items() if "paid" in k.lower())
        total_pending = sum(v for k, v in summary["financial_summary"].items() if "pending" in k.lower())
        total_rejected = sum(v for k, v in summary["financial_summary"].items() if "rejected" in k.lower())
        total_petty = sum(v for k, v in summary["financial_summary"].items() if "petty" in k.lower())
        
        report_text += f"  • Total Paid: RM {total_paid:,.2f}\n"
        report_text += f"  • Total Pending/KIV: RM {total_pending:,.2f}\n"
        report_text += f"  • Petty Cash Claims: RM {total_petty:,.2f}\n"
        report_text += f"  • Rejected Claims: RM {total_rejected:,.2f}\n"
        
        grand_total = total_paid + total_pending + total_petty
        report_text += f"\n  🎯 TOTAL VALID CLAIMS (Paid + Pending + Petty): RM {grand_total:,.2f}\n"
        
        report_text += f"\n✅ CLEANED DATA FILE: {cleaned_filename}\n"
        report_text += f"📊 Ready for download and further analysis\n"
        
        # Use LLM to generate business insights from cleaned data
        insights_prompt = f"""
Based on this cleaned financial data summary:

{report_text}

Provide brief business insights about:
1. Financial health status (what's paid vs pending)
2. Risk areas (high rejected amount, pending claims concerns)
3. Cash flow implications
4. Recommendations for follow-up

Keep it concise and actionable.
"""
        
        insights = llm.invoke(insights_prompt)
        
        return {
            "status": "success",
            "file_name": file_name,
            "cleaned_file": cleaned_filename,
            "report": report_text,
            "business_insights": insights.content,
            "summary": {
                "total_paid": round(total_paid, 2),
                "total_pending": round(total_pending, 2),
                "total_petty": round(total_petty, 2),
                "total_rejected": round(total_rejected, 2),
                "grand_total": round(grand_total, 2),
                "status_breakdown": summary["status_breakdown"],
            }
        }
    
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to clean and analyze data: {str(e)}",
        }


async def generate_insights(file_path: str, file_name: str) -> Dict[str, Any]:
    """Analyze a file (data or PDF) and generate intelligent insights using LLM."""
    try:
        # Determine file type and extract statistics
        if file_name.lower().endswith(".csv"):
            stats = _analyze_csv(file_path)
            is_pdf = False
        elif file_name.lower().endswith((".xlsx", ".xls")):
            stats = _analyze_excel(file_path)
            is_pdf = False
        elif file_name.lower().endswith(".pdf"):
            stats = _analyze_pdf(file_path)
            is_pdf = True
        else:
            return {
                "status": "error",
                "message": f"Unsupported file type for analysis: {file_name}",
            }
        
        # Handle PDF text analysis
        if is_pdf:
            stats_text = f"""
**DOCUMENT ANALYSIS: {file_name}**

Total Pages: {stats['total_pages']}
Total Words: {stats['total_words']:,}
Total Sentences: {stats['total_sentences']}
Average Page Length: {stats['average_page_length']:.0f} characters

**KEY TOPICS (by word frequency)**:
"""
            for word, freq in list(stats['word_frequency'].items())[:10]:
                stats_text += f"  - {word}: {freq} occurrences\n"
            
            stats_text += f"\n**DOCUMENT PREVIEW**:\n{stats['full_text'][:1500]}\n...\n"
            
            # Use LLM for PDF document analysis
            prompt = f"""
You are a expert senior project management consultant for a construction company. Analyze this PDF document and provide INTELLIGENT, ACTIONABLE insights.

{stats_text}

Based on the document content and structure, provide:

1. **DOCUMENT OVERVIEW**: What is the purpose and scope of this document?

2. **KEY FINDINGS & INSIGHTS**: What are the most important points, achievements, or concerns discussed?

3. **PROJECT STATUS**: Extract and summarize:
   - Current project status and progress
   - Key milestones achieved or pending
   - Budget/cost implications if mentioned
   - Timeline and schedule status

4. **RISKS & ISSUES**: Identify:
   - Any mentioned risks or challenges
   - Outstanding issues or blockers
   - Recommended mitigation strategies

5. **ACTION ITEMS**: What specific actions or decisions are recommended based on this document?

Focus on BUSINESS IMPACT and provide specific references to document content.
"""
        else:
            # Original data analysis logic
            stats_text = f"""
**DATASET OVERVIEW**
File: {file_name}
Rows: {stats['total_rows']:,} | Columns: {stats['total_columns']}
Data Quality Score: {stats['data_quality_score']}% | Duplicate Records: {stats['duplicates']}

**COLUMNS**: {', '.join(stats['column_names'])}

**DATA COMPLETENESS ISSUES**:
"""
            
            # Missing value details
            significant_missing = {k: v for k, v in stats['missing_percentage'].items() if v > 5}
            if significant_missing:
                stats_text += "\n".join([f"  - {col}: {v}% missing" for col, v in significant_missing.items()])
            else:
                stats_text += "  - No significant missing values (all columns < 5% missing)"
            
            # Numeric analysis
            if "numeric_statistics" in stats:
                stats_text += "\n\n**NUMERIC COLUMNS ANALYSIS**:\n"
                for col, col_stats in stats["numeric_statistics"].items():
                    stats_text += f"\n  {col}:\n"
                    stats_text += f"    - Range: {col_stats['min']} to {col_stats['max']}\n"
                    stats_text += f"    - Mean: {col_stats['mean']}, Median: {col_stats['median']}, Std Dev: {col_stats['std']}\n"
                    stats_text += f"    - IQR: {col_stats['q25']} to {col_stats['q75']}\n"
                    if 'outliers' in col_stats:
                        stats_text += f"    - ⚠️ OUTLIERS DETECTED: {col_stats['outliers']['count']} records ({col_stats['outliers']['percentage']}%) with unusual values\n"
            
            # Categorical analysis
            if "categorical_summary" in stats:
                stats_text += "\n**CATEGORICAL COLUMNS ANALYSIS**:\n"
                for col, col_data in stats["categorical_summary"].items():
                    stats_text += f"\n  {col}:\n"
                    stats_text += f"    - Unique values: {col_data['unique_values']}\n"
                    stats_text += f"    - Most common: {col_data['most_common']}\n"
                    stats_text += f"    - Distribution: {col_data['value_counts']}\n"
            
            # Correlations
            if "strong_correlations" in stats:
                stats_text += "\n**STRONG CORRELATIONS** (r > 0.7):\n"
                for corr in stats["strong_correlations"]:
                    stats_text += f"  - {corr['var1']} ↔ {corr['var2']}: {corr['correlation']}\n"
            
            # Use LLM to generate intelligent insights
            prompt = f"""
You are a expert senior data analyst for a construction/project management company. Analyze this dataset and provide INTELLIGENT, ACTIONABLE business insights.

{stats_text}

Based on this analysis, provide:

1. **DATA QUALITY ASSESSMENT**: Is this data reliable? What are the risks?

2. **KEY FINDINGS**: What patterns, anomalies, or important trends do you see?

3. **CRITICAL ISSUES**: Highlight any data problems that could impact decision-making:
   - Data inconsistencies or quality issues
   - Missing values that could skew results
   - Outliers that suggest errors or exceptional cases
   - Data that needs cleaning

4. **BUSINESS INSIGHTS**: What does this data tell us about:
   - Performance metrics
   - Resource utilization
   - Cost/budget implications
   - Risk factors
   - Operational efficiency

5. **RECOMMENDATIONS**: What actions should be taken based on these findings?

Be specific, use the actual numbers from the data, and focus on BUSINESS IMPACT rather than technical details.
Keep insights concise but substantive.
"""
        
        response = llm.invoke(prompt)
        
        # Build response based on file type
        if is_pdf:
            return {
                "status": "success",
                "file_name": file_name,
                "file_type": "PDF Document",
                "total_pages": stats['total_pages'],
                "total_words": stats['total_words'],
                "insights": response.content,
                "statistics": stats,
            }
        else:
            return {
                "status": "success",
                "file_name": file_name,
                "file_type": "Data File",
                "total_rows": stats['total_rows'],
                "total_columns": stats['total_columns'],
                "data_quality": stats['data_quality_score'],
                "insights": response.content,
                "statistics": stats,
            }
    
    except Exception as e:
        return {
            "status": "error",
            "message": f"Failed to analyze file: {str(e)}",
        }
